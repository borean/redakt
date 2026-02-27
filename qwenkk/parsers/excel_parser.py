import shutil
from pathlib import Path

from openpyxl import load_workbook

from qwenkk.core.entities import PIIEntity
from qwenkk.parsers.base import BaseParser, ParseResult


class ExcelParser(BaseParser):

    def extract_text(self, file_path: Path) -> ParseResult:
        wb = load_workbook(str(file_path), data_only=True)
        all_text: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        all_text.append(str(cell.value))

        wb.close()
        return ParseResult(
            text="\n".join(all_text),
            file_path=file_path,
            metadata={"sheets": len(wb.sheetnames)},
        )

    def create_anonymized(
        self,
        file_path: Path,
        entities: list[PIIEntity],
        output_path: Path | None = None,
    ) -> Path:
        output_path = output_path or self.output_filename(file_path)
        shutil.copy2(file_path, output_path)

        wb = load_workbook(str(output_path))
        # Sort by length descending
        sorted_replacements = sorted(
            [(e.original, e.placeholder) for e in entities],
            key=lambda x: len(x[0]),
            reverse=True,
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue
                    text = cell.value
                    for orig, placeholder in sorted_replacements:
                        text = text.replace(orig, placeholder)
                    cell.value = text

        wb.save(str(output_path))
        wb.close()
        return output_path
